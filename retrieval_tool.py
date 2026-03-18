import argparse
from decimal import Decimal, InvalidOperation
import hashlib
import hmac
import json
import os
from pathlib import Path
import re
import shutil
import sys
import tempfile
from typing import Any, Callable, Dict, List, Optional, Set, Tuple

KEY_COLUMN_A = "/A"
KEY_COLUMN_B = "/B"
KEY_COLUMN_C = "/C"
KEY_COLUMN_D = "/D"
KEY_COLUMN_E = "/E"
KEY_COLUMN_F = "/F"
ALL_KEY_COLUMNS = {KEY_COLUMN_A, KEY_COLUMN_B, KEY_COLUMN_C, KEY_COLUMN_D, KEY_COLUMN_E, KEY_COLUMN_F}
FUZZY_KEY_PAIRS = ((KEY_COLUMN_C, KEY_COLUMN_D), (KEY_COLUMN_E, KEY_COLUMN_F))
HIDDEN_RESULT_HEADERS = set(ALL_KEY_COLUMNS)
INITIAL_HEADERS = ["序号", KEY_COLUMN_A, "直径", "长度", KEY_COLUMN_B]
SETTINGS_PASSCODE = "2233"
MAIN_WINDOW_SIZE = "900x560"
MAIN_WINDOW_TOP_OFFSET = 40
SETTINGS_WINDOW_WIDTH = 760
SETTINGS_WINDOW_HEIGHT = 420
TABLE_COLUMN_WIDTH = 120
QUERY_INPUT_WIDTH = 12
TYPE_ICON_TARGET_SIZE = 240
DEFERRED_STARTUP_ICON_DELAY_MS = 120
AUTO_FIT_SAMPLE_LIMIT = 200
APP_ICON_FILE = "logo.png"
OPTIONS_ICON_FILE = "options.png"
TYPES_CONFIG_FILE = "types_config.json"
UI_CONFIG_FILE = "ui_config.json"
DEFAULT_MAIN_TITLE = "标题"
APP_VERSION = "0.7.260319"
APP_FOOTER_TEXT = f"Powered by GPT & ZL | v{APP_VERSION}"
# types_config.json 新格式字段：
# - _data: 实际类型配置
# - _signature: 用口令计算出的签名，用于检测文件是否被手动改动
CONFIG_DATA_FIELD = "_data"
CONFIG_SIGNATURE_FIELD = "_signature"


def save_settings_atomically(
    save_title_func: Callable[[], bool],
    save_mapping_func: Callable[[], bool],
    rollback_title_func: Callable[[], bool],
) -> bool:
    """Save settings all-or-nothing: title first, mapping second, title rollback on failure."""
    if not save_title_func():
        return False

    if save_mapping_func():
        return True

    rollback_title_func()
    return False


def get_openpyxl_symbols():
    # 懒加载 openpyxl，减少程序启动阶段的导入开销
    from openpyxl import Workbook, load_workbook

    return Workbook, load_workbook


def get_app_base_dir() -> Path:
    """Return writable app directory (script dir in dev, exe dir when frozen)."""
    # 开发模式：写在源码目录。
    # 打包后：写在 exe 所在目录，方便同事直接使用。
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def get_resource_path(file_name: str) -> Path:
    """Resolve bundled resource path for both source and PyInstaller modes."""
    # PyInstaller 单文件模式会把资源解包到临时目录(_MEIPASS)，
    # 这里优先找临时目录中的资源；找不到再回退到程序目录。
    if getattr(sys, "frozen", False):
        bundle_dir = Path(getattr(sys, "_MEIPASS", get_app_base_dir()))
        bundled_file = bundle_dir / file_name
        if bundled_file.exists():
            return bundled_file
    return get_app_base_dir() / file_name


def ensure_config_file(config_path: Path) -> None:
    """Ensure writable config exists next to app, copying bundled default if available."""
    if config_path.exists():
        return

    bundled_config = get_resource_path(TYPES_CONFIG_FILE)
    if bundled_config.exists() and bundled_config != config_path:
        try:
            shutil.copy2(bundled_config, config_path)
        except OSError:
            pass


def get_user_config_dir() -> Path:
    """Return per-user writable directory for app config."""
    if sys.platform.startswith("win"):
        base_dir = Path.home()
        local_app_data = os.environ.get("LOCALAPPDATA")
        if local_app_data:
            base_dir = Path(local_app_data)
        return base_dir / "RetrievalTool"

    return Path.home() / ".config" / "retrieval_tool"


def get_user_config_path() -> Path:
    return get_user_config_dir() / TYPES_CONFIG_FILE


def get_preferred_config_path() -> Path:
    """Pick app dir config when possible, otherwise fallback to user-writable path."""
    # 开发模式优先用用户目录，避免运行时改动仓库内的配置文件。
    if not getattr(sys, "frozen", False):
        try:
            user_config_dir = get_user_config_dir()
            user_config_dir.mkdir(parents=True, exist_ok=True)
            user_config_path = get_user_config_path()
            ensure_config_file(user_config_path)
            if not user_config_path.exists():
                user_config_path.write_text(dump_signed_type_mapping({}), encoding="utf-8")
            return user_config_path
        except OSError:
            pass

    app_config_path = get_app_base_dir() / TYPES_CONFIG_FILE
    ensure_config_file(app_config_path)
    if app_config_path.exists():
        return app_config_path

    try:
        user_config_dir = get_user_config_dir()
        user_config_dir.mkdir(parents=True, exist_ok=True)
        user_config_path = get_user_config_path()
        ensure_config_file(user_config_path)
        if not user_config_path.exists():
            user_config_path.write_text(dump_signed_type_mapping({}), encoding="utf-8")
        return user_config_path
    except OSError:
        # 用户目录不可写时回退到程序目录（可能不存在，后续按空配置处理）。
        return app_config_path


def get_query_headers(headers: List[str]) -> List[str]:
    # 可查询字段定义为“/A”和“/B”之间的所有列
    """Return queryable headers between /A and /B."""
    key_columns = resolve_key_columns(headers)
    if key_columns is None:
        return []

    key_a, key_b = key_columns
    key_a_idx = headers.index(key_a)
    key_b_idx = headers.index(key_b)
    if key_a_idx >= key_b_idx:
        return []

    return [header for header in headers[key_a_idx + 1 : key_b_idx] if header not in ALL_KEY_COLUMNS]


def resolve_key_columns(headers: List[str]) -> Optional[Tuple[str, str]]:
    """Return active key column pair."""
    if KEY_COLUMN_A in headers and KEY_COLUMN_B in headers:
        return KEY_COLUMN_A, KEY_COLUMN_B
    return None


def get_fuzzy_query_fields(headers: List[str]) -> Set[str]:
    """Return fields that should use fuzzy (contains) matching."""
    key_columns = resolve_key_columns(headers)
    if key_columns is None:
        return set()

    key_a, key_b = key_columns
    key_a_idx = headers.index(key_a)
    key_b_idx = headers.index(key_b)
    fuzzy_fields: Set[str] = set()
    present_ranges: List[Tuple[str, str, int, int]] = []

    for start_key, end_key in FUZZY_KEY_PAIRS:
        has_start = start_key in headers
        has_end = end_key in headers
        if has_start != has_end:
            raise ValueError(f"标题行关键列不完整：{start_key} 与 {end_key} 必须成对出现。")
        if not has_start:
            continue

        start_idx = headers.index(start_key)
        end_idx = headers.index(end_key)
        if start_idx >= end_idx:
            raise ValueError(f"标题顺序错误：必须先有“{start_key}”，后有“{end_key}”。")
        if not (key_a_idx < start_idx < end_idx < key_b_idx):
            raise ValueError(
                f"标题位置错误：{start_key} 与 {end_key} 必须位于“{KEY_COLUMN_A}”和“{KEY_COLUMN_B}”之间。"
            )

        present_ranges.append((start_key, end_key, start_idx, end_idx))

        fuzzy_fields.update(header for header in headers[start_idx + 1 : end_idx] if header not in ALL_KEY_COLUMNS)

    if len(present_ranges) >= 2:
        _, _, start1, end1 = present_ranges[0]
        _, _, start2, end2 = present_ranges[1]
        is_cross = (start1 < start2 < end1 < end2) or (start2 < start1 < end2 < end1)
        if is_cross:
            raise ValueError(
                f"标题格式错误：{KEY_COLUMN_C}/{KEY_COLUMN_D} 与 {KEY_COLUMN_E}/{KEY_COLUMN_F} 不能交叉。"
            )

    return fuzzy_fields


def get_display_headers(headers: List[str]) -> List[str]:
    """Return headers visible in result output/table."""
    return [header for header in headers if header not in HIDDEN_RESULT_HEADERS]


def load_type_mapping_from_file_path(config_path: Path) -> Dict[str, Dict[str, str]]:
    """Load type mapping with backward compatibility and safe fallback."""
    mapping, warning_message = load_type_mapping_with_status(config_path)
    if warning_message:
        print(f"[WARN] {warning_message}")
    return mapping


def load_type_mapping_with_status(config_path: Path) -> Tuple[Dict[str, Dict[str, str]], Optional[str]]:
    """Load type mapping and return a user-facing warning when config is rejected."""
    if not config_path.exists():
        return {}, None

    try:
        raw_config = json.loads(config_path.read_text(encoding="utf-8"))

        if not isinstance(raw_config, dict):
            return {}, "类型配置格式错误，已回退为空配置。"

        # 新版受保护格式：必须同时有 _data 和 _signature。
        # 少任何一个都视为手工改动（或损坏），拒绝加载。
        has_data_field = CONFIG_DATA_FIELD in raw_config
        has_signature_field = CONFIG_SIGNATURE_FIELD in raw_config
        if has_data_field or has_signature_field:
            # If one of the protected fields is missing, treat as tampering/corruption.
            if not (has_data_field and has_signature_field):
                return {}, "检测到类型配置被手动修改（签名字段缺失），已忽略该配置。"

            raw_data = raw_config.get(CONFIG_DATA_FIELD, {})
            raw_signature = str(raw_config.get(CONFIG_SIGNATURE_FIELD, ""))
            expected_signature = sign_type_mapping(raw_data if isinstance(raw_data, dict) else {})
            if not hmac.compare_digest(raw_signature, expected_signature):
                return {}, "检测到类型配置被手动修改（签名校验失败），已忽略该配置。"
            config_data = raw_data
        else:
            # 旧版兼容：只允许 {"名称": "excel路径"} 这种纯字符串格式。
            # 如果是未签名的新结构（value 是 dict），也拒绝加载，避免绕过校验。
            if any(not isinstance(value, str) for value in raw_config.values()):
                return {}, "检测到类型配置被手动修改（未签名结构），已忽略该配置。"
            config_data = raw_config

        cleaned_mapping: Dict[str, Dict[str, str]] = {}
        for key, value in config_data.items():
            key_text = str(key).strip()
            if not key_text:
                continue

            # Backward compatible: old format is {"名称": "excel路径"}
            if isinstance(value, str):
                value_text = value.strip()
                if value_text:
                    cleaned_mapping[key_text] = {"excel": value_text, "icon": ""}
                continue

            if not isinstance(value, dict):
                continue

            excel_text = str(value.get("excel", "")).strip()
            icon_text = str(value.get("icon", "")).strip()
            if excel_text:
                cleaned_mapping[key_text] = {"excel": excel_text, "icon": icon_text}
        return cleaned_mapping, None
    except (json.JSONDecodeError, OSError, TypeError, ValueError) as exc:
        return {}, f"类型配置读取失败，已回退为空配置: {exc}"


def sign_type_mapping(mapping: Dict[str, Dict[str, str]]) -> str:
    """Create an HMAC signature for config integrity validation."""
    # 注意：这里是“完整性检测”，不是强安全防攻击方案。
    # 目标是提示误改/误删，不是防止有能力的逆向者伪造。
    payload = json.dumps(mapping, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    key = SETTINGS_PASSCODE.encode("utf-8")
    return hmac.new(key, payload, hashlib.sha256).hexdigest()


def dump_signed_type_mapping(mapping: Dict[str, Dict[str, str]]) -> str:
    """Serialize type mapping with signature so tampering can be detected."""
    signed_payload = {
        CONFIG_DATA_FIELD: mapping,
        CONFIG_SIGNATURE_FIELD: sign_type_mapping(mapping),
    }
    return json.dumps(signed_payload, ensure_ascii=False, indent=2)


def get_icon_subsample_scale(width: int, height: int, target_size: int) -> int:
    """Return integer subsample scale so output does not exceed target size."""
    # Tk 的 subsample 只能用整数倍率，
    # 这里通过“向上取整”保证缩放后不会超过目标尺寸。
    safe_width = max(width, 1)
    safe_height = max(height, 1)
    safe_target = max(target_size, 1)
    ratio = max(safe_width / safe_target, safe_height / safe_target)
    scale = int(ratio)
    if ratio > scale:
        scale += 1
    return max(scale, 1)


def normalize_value(value) -> str:
    # 统一查询值与表格值的比较口径，减少格式差异导致的误判
    """Convert worksheet/user input values to comparable text."""
    if value is None:
        return ""

    text = str(value).strip()
    if not text:
        return ""

    # /A 这类值可能有前导 0（如 037332314020），不能被当数字吞掉 0。
    if text.isdigit() and len(text) > 1 and text.startswith("0"):
        return text

    # 其余“看起来像数字”的文本，统一成可比较格式：
    # 例如 1200.50 -> 1200.5，便于查询时匹配。
    try:
        number = Decimal(text)
    except InvalidOperation:
        return text

    if number == number.to_integral_value():
        integer_text = str(number.quantize(Decimal(1)))
        return "0" if integer_text in {"-0", "+0"} else integer_text

    normalized = format(number.normalize(), "f")
    if "." in normalized:
        normalized = normalized.rstrip("0").rstrip(".")
    if normalized in {"-0", "+0", ""}:
        return "0"
    return normalized


def normalize_match_key(value) -> str:
    """Normalize value for query matching (case-insensitive)."""
    return normalize_value(value).casefold()


def format_cell_value(cell) -> str:
    # 优先按单元格显示规则取值，避免关键字段丢失前导 0
    """Format an openpyxl cell value while preserving zero-padded number formats."""
    value = cell.value
    if value is None:
        return ""

    if isinstance(value, bool):
        return str(value)

    if isinstance(value, (int, float)):
        if isinstance(value, float) and not value.is_integer():
            return normalize_value(value)

        int_value = int(value)
        number_format = str(cell.number_format or "").strip()
        # For formats like 000000000000, preserve leading zeros in display text.
        if re.fullmatch(r"0+", number_format):
            return f"{int_value:0{len(number_format)}d}"

        return str(int_value)

    return normalize_value(value)


def load_records(excel_path: Path) -> Tuple[List[str], List[Dict[str, str]]]:
    # 读取首行为动态表头，并按表头映射后续行到字典
    if not excel_path.exists():
        raise FileNotFoundError(f"未找到 Excel 文件: {excel_path}")

    _, load_workbook = get_openpyxl_symbols()
    workbook = load_workbook(filename=excel_path, data_only=True, read_only=True)
    try:
        worksheet = workbook.active

        first_row = next(
            worksheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=worksheet.max_column, values_only=False),
            None,
        )
        if first_row is None:
            raise ValueError("Excel 文件为空，缺少标题行。")

        actual_headers = [normalize_value(cell.value) for cell in first_row]
        while actual_headers and not actual_headers[-1]:
            actual_headers.pop()

        if not actual_headers:
            raise ValueError("Excel 文件缺少有效标题行。")

        # 标题行中间不能留空列，否则后续字段映射会错位。
        if any(not header for header in actual_headers):
            raise ValueError(
                "标题行不完整，请确保标题之间没有空列。\n"
                f"实际: {actual_headers}"
            )

        key_columns = resolve_key_columns(actual_headers)
        if key_columns is None:
            raise ValueError(f"标题行缺少必需列: {KEY_COLUMN_A}/{KEY_COLUMN_B}")

        key_a, key_b = key_columns
        if actual_headers.index(key_a) >= actual_headers.index(key_b):
            raise ValueError(f"标题顺序错误：必须先有“{key_a}”，后有“{key_b}”。")

        # 校验 /C-/D、/E-/F 模糊区间关键列的合法性。
        get_fuzzy_query_fields(actual_headers)

        # 把每一行转成 {标题: 值} 的字典，后续查询统一按字典处理。
        records: List[Dict[str, str]] = []
        for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=len(actual_headers), values_only=False):
            if row is None or all(cell.value is None for cell in row):
                continue

            record: Dict[str, str] = {}
            for idx, header in enumerate(actual_headers):
                record[header] = format_cell_value(row[idx]) if idx < len(row) else ""
            records.append(record)

        return actual_headers, records
    finally:
        workbook.close()


def query_records(
    records: List[Dict[str, str]],
    conditions: Dict[str, str],
    candidate_indexes: Optional[List[int]] = None,
    fuzzy_fields: Optional[Set[str]] = None,
    normalized_rows: Optional[List[Dict[str, str]]] = None,
) -> List[Dict[str, str]]:
    # 仅保留有效查询条件，空输入不参与过滤
    normalized_conditions: Dict[str, str] = {}
    for key, value in conditions.items():
        normalized = normalize_match_key(value)
        if normalized:
            normalized_conditions[key] = normalized

    if not normalized_conditions:
        return []

    normalized_cache_ready = normalized_rows is not None and len(normalized_rows) == len(records)
    if candidate_indexes is None:
        if normalized_cache_ready:
            candidate_pairs = zip(records, normalized_rows)
        else:
            candidate_pairs = ((row, None) for row in records)
    else:
        record_count = len(records)
        if normalized_cache_ready:
            candidate_pairs = (
                (records[idx], normalized_rows[idx])
                for idx in candidate_indexes
                if 0 <= idx < record_count
            )
        else:
            candidate_pairs = ((records[idx], None) for idx in candidate_indexes if 0 <= idx < record_count)

    fuzzy_match_fields = fuzzy_fields or set()

    def matches(row: Dict[str, str], normalized_row: Optional[Dict[str, str]]) -> bool:
        # 多条件“且”匹配：任一条件不满足即排除
        for key, value in normalized_conditions.items():
            if normalized_row is not None and key in normalized_row:
                row_value = normalized_row.get(key, "")
            else:
                row_value = normalize_match_key(row.get(key, ""))
            if key in fuzzy_match_fields:
                if value not in row_value:
                    return False
            elif row_value != value:
                return False
        return True

    return [row for row, normalized_row in candidate_pairs if matches(row, normalized_row)]


def build_normalized_query_rows(records: List[Dict[str, str]], fields: List[str]) -> List[Dict[str, str]]:
    """Build normalized values cache for query fields to speed up repeated filtering."""
    if not records or not fields:
        return [{} for _ in records]

    normalized_rows: List[Dict[str, str]] = []
    for row in records:
        normalized_rows.append({field: normalize_match_key(row.get(field, "")) for field in fields})
    return normalized_rows


def build_query_indexes(records: List[Dict[str, str]], fields: List[str]) -> Dict[str, Dict[str, List[int]]]:
    """Build in-memory indexes for faster repeated exact-match queries."""
    indexes: Dict[str, Dict[str, List[int]]] = {field: {} for field in fields}
    for idx, row in enumerate(records):
        for field in fields:
            normalized = normalize_match_key(row.get(field))
            if not normalized:
                continue
            indexes[field].setdefault(normalized, []).append(idx)
    return indexes


def get_candidate_indexes(
    indexes: Dict[str, Dict[str, List[int]]],
    conditions: Dict[str, str],
) -> Optional[List[int]]:
    """Return ordered candidate row indexes by intersecting indexed conditions."""
    # 思路：
    # 1) 先从每个条件拿到命中的行号列表
    # 2) 再做交集，得到“同时满足所有条件”的候选行
    indexed_lists: List[List[int]] = []
    for field, raw_value in conditions.items():
        normalized = normalize_match_key(raw_value)
        if not normalized:
            continue
        if field not in indexes:
            continue

        row_ids = indexes[field].get(normalized)
        if not row_ids:
            return []
        indexed_lists.append(row_ids)

    if not indexed_lists:
        return None

    # 从最短列表开始过滤，能减少交集计算成本。
    base = min(indexed_lists, key=len)
    other_sets = [set(ids) for ids in indexed_lists if ids is not base]
    if not other_sets:
        return list(base)

    return [row_id for row_id in base if all(row_id in each_set for each_set in other_sets)]


def run_self_check() -> bool:
    """Development-only lightweight checks for critical data handling paths."""
    try:
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            Workbook, _ = get_openpyxl_symbols()

            # 1) 前导 0 关键字段保留验证
            excel_path = tmp_path / "self_check.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["序号", KEY_COLUMN_A, "直径", "长度", KEY_COLUMN_B])

            code_cell = ws.cell(row=2, column=2, value=37332314020)
            code_cell.number_format = "000000000000"
            ws.cell(row=2, column=1, value=1)
            ws.cell(row=2, column=3, value=20)
            ws.cell(row=2, column=4, value=1200.5)
            ws.cell(row=2, column=5, value="项目A")
            wb.save(excel_path)
            wb.close()

            headers, records = load_records(excel_path)
            if not records or records[0].get(KEY_COLUMN_A) != "037332314020":
                print("[FAIL] 前导0关键字段保留校验失败")
                return False

            # 2) 小数查询条件验证
            decimal_results = query_records(records, {"长度": "1200.50"})
            if len(decimal_results) != 1:
                print("[FAIL] 小数查询校验失败")
                return False

            # 3) 配置文件损坏回退验证
            config_path = tmp_path / "types_config.json"
            config_path.write_text("{invalid_json", encoding="utf-8")
            config = load_type_mapping_from_file_path(config_path)
            if config != {}:
                print("[FAIL] 损坏配置回退校验失败")
                return False

            print("[PASS] 自检通过：前导0、小数查询、损坏配置回退")
            print(f"[INFO] 测试表头: {headers}")
            return True
    except Exception as exc:
        print(f"[FAIL] 自检异常: {exc}")
        return False


def print_results(headers: List[str], results: List[Dict[str, str]]) -> None:
    if not results:
        print("未查询到匹配数据。")
        return

    visible_headers = get_display_headers(headers)
    print(f"查询到 {len(results)} 条结果:")
    print("-" * 50)
    for item in results:
        line = " | ".join(f"{header}: {item.get(header, '')}" for header in visible_headers)
        print(line)


def run_interactive(headers: List[str], records: List[Dict[str, str]]) -> None:
    # 命令行模式使用与 GUI 一致的动态查询字段
    print("已加载数据，可开始查询。输入 q 退出。")

    query_headers = get_query_headers(headers)
    fuzzy_fields = get_fuzzy_query_fields(headers)
    if not query_headers:
        print(f"当前表格在“{KEY_COLUMN_A}”和“{KEY_COLUMN_B}”之间没有可查询字段。")
        return

    while True:
        conditions: Dict[str, str] = {}
        for header in query_headers:
            value = input(f"请输入{header}（可留空）: ").strip()
            if value.lower() == "q":
                return
            if value:
                conditions[header] = value

        if not conditions:
            print("请至少输入一个查询条件。")
            print("-" * 50)
            continue

        results = query_records(records, conditions, fuzzy_fields=fuzzy_fields)
        print_results(headers, results)
        print("-" * 50)


def launch_gui(default_excel: Path, prefer_default_excel: bool = False) -> None:
    import tkinter as tk
    import tkinter.font as tkfont
    from tkinter import filedialog, messagebox, simpledialog, ttk

    # GUI 主入口：类型管理、动态查询项、结果展示与复制能力
    root = tk.Tk()
    root.withdraw()
    root.title("检索工具")
    root.geometry(MAIN_WINDOW_SIZE)
    # 先隐藏窗口，算好最终位置后再显示，避免“先在别处闪一下”。
    root.update_idletasks()

    # 启动时窗口居中并略微上移，避免视觉重心过低。
    window_w = root.winfo_width()
    window_h = root.winfo_height()
    screen_w = root.winfo_screenwidth()
    screen_h = root.winfo_screenheight()
    pos_x = max((screen_w - window_w) // 2, 0)
    pos_y = max((screen_h - window_h) // 2 - MAIN_WINDOW_TOP_OFFSET, 0)
    root.geometry(f"{window_w}x{window_h}+{pos_x}+{pos_y}")
    root.deiconify()

    # 程序标题栏图标：优先提供多尺寸（含 256x256），让系统自动选择最合适图标。
    # 为了提升首屏速度，图标加载放到窗口空闲时执行。
    app_icon_images: List[Any] = []
    icon_path = get_resource_path(APP_ICON_FILE)
    options_icon_path = get_resource_path(OPTIONS_ICON_FILE)

    def apply_window_icon() -> None:
        nonlocal app_icon_images
        if not icon_path.exists():
            return

        try:
            from PIL import Image, ImageTk

            with Image.open(icon_path) as icon_source:
                if icon_source.mode not in ("RGB", "RGBA"):
                    icon_source = icon_source.convert("RGBA")

                icon_sizes = [16, 32, 48, 64, 128, 256]
                for size in icon_sizes:
                    resized = icon_source.resize((size, size), Image.Resampling.LANCZOS)
                    app_icon_images.append(ImageTk.PhotoImage(resized))

            if app_icon_images:
                root.iconphoto(True, *app_icon_images)
        except (ImportError, OSError, ValueError, tk.TclError):
            try:
                fallback_icon = tk.PhotoImage(file=str(icon_path))
                app_icon_images = [fallback_icon]
                root.iconphoto(True, fallback_icon)
            except tk.TclError:
                app_icon_images = []

    style = ttk.Style(root)
    default_font = tkfont.nametofont("TkDefaultFont")
    table_measure_font = tkfont.nametofont("TkDefaultFont")
    title_font = tkfont.Font(
        root=root,
        family="华文隶书",
        size=max(int(default_font.cget("size")) * 3, 1),
        weight="bold",
    )
    query_font = tkfont.Font(root=root, family=default_font.cget("family"), size=default_font.cget("size"), weight="bold")
    style.configure("Query.TButton", font=query_font, foreground="#000000")

    # records: 当前已加载 Excel 的原始记录
    # query_indexes: 为查询字段建立的倒排索引（加速查询）
    records: List[Dict[str, str]] = []
    query_indexes: Dict[str, Dict[str, List[int]]] = {}
    normalized_query_rows: List[Dict[str, str]] = []
    current_headers: List[str] = []
    sort_descending: Dict[str, bool] = {}
    current_sort_column: Optional[str] = None
    current_sort_is_desc: bool = False
    query_vars: Dict[str, tk.StringVar] = {}
    fuzzy_query_fields: Set[str] = set()
    config_path: Optional[Path] = None
    type_mapping: Dict[str, Dict[str, str]] = {}
    type_var = tk.StringVar()
    app_title_var = tk.StringVar(value=DEFAULT_MAIN_TITLE)
    status_var = tk.StringVar(value="正在初始化配置...")
    type_icon_image: Optional[Any] = None
    icon_cache: Dict[Tuple[str, int], Any] = {}
    pending_icon_update_id: Optional[str] = None
    pending_full_autofit_id: Optional[str] = None
    render_generation = 0
    settings_button_icon_image: Optional[Any] = None
    startup_load_attempted = False
    # loaded_type_name 表示“真正加载成功”的类型。
    # 仅切换下拉框不算加载，避免图标/状态误导用户。
    loaded_type_name = ""
    default_type_icon_path = str(icon_path) if icon_path.exists() else ""

    root.columnconfigure(0, weight=1)
    root.rowconfigure(2, weight=1)

    top_frame = ttk.Frame(root, padding=10)
    top_frame.grid(row=0, column=0, sticky="ew")
    top_frame.columnconfigure(1, weight=1)

    type_icon_canvas = tk.Canvas(
        top_frame,
        width=TYPE_ICON_TARGET_SIZE,
        height=TYPE_ICON_TARGET_SIZE,
        highlightthickness=0,
        bd=0,
        background=root.cget("bg"),
    )
    type_icon_canvas.grid(row=0, column=0, sticky="nw", padx=(18, 28), pady=(0, 0))

    type_control_frame = ttk.Frame(top_frame)
    type_control_frame.grid(row=0, column=1, sticky="ew", pady=(10, 0))
    type_control_frame.columnconfigure(1, weight=1)
    ttk.Label(type_control_frame, textvariable=app_title_var, font=title_font, anchor="center").grid(
        row=0,
        column=0,
        columnspan=3,
        sticky="ew",
        pady=(0, 12),
    )
    ttk.Label(type_control_frame, text="类型:").grid(row=1, column=0, sticky="w", padx=(0, 8))
    type_combo = ttk.Combobox(type_control_frame, textvariable=type_var, state="readonly", width=56)
    type_combo.grid(row=1, column=1, sticky="ew")

    query_frame = ttk.Frame(root, padding=(10, 0, 10, 10))
    query_frame.grid(row=1, column=0, sticky="ew")
    query_frame.columnconfigure(0, weight=1)
    query_input_frame = ttk.Frame(query_frame)
    query_input_frame.grid(row=0, column=0, sticky="ew")

    table_frame = ttk.Frame(root, padding=(10, 0, 10, 10))
    table_frame.grid(row=2, column=0, sticky="nsew")
    table_frame.rowconfigure(0, weight=1)
    table_frame.columnconfigure(0, weight=1)

    columns = tuple(current_headers)
    table = ttk.Treeview(table_frame, columns=columns, show="headings")
    for col in columns:
        table.heading(col, text=col, command=lambda c=col: sort_table_by_column(c))
        table.column(col, width=TABLE_COLUMN_WIDTH, anchor="center")

    table.grid(row=0, column=0, sticky="nsew")

    y_scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=table.yview)
    y_scrollbar.grid(row=0, column=1, sticky="ns")
    x_scrollbar = ttk.Scrollbar(table_frame, orient="horizontal", command=table.xview)
    x_scrollbar.grid(row=1, column=0, sticky="ew")
    x_scrollbar_visible = True

    def update_x_scrollbar(first, last) -> None:
        nonlocal x_scrollbar_visible
        x_scrollbar.set(first, last)
        try:
            first_f = float(first)
            last_f = float(last)
        except (TypeError, ValueError):
            return

        needs_scroll = first_f > 0.0 or last_f < 1.0
        if needs_scroll and not x_scrollbar_visible:
            x_scrollbar.grid()
            x_scrollbar_visible = True
        elif not needs_scroll and x_scrollbar_visible:
            x_scrollbar.grid_remove()
            x_scrollbar_visible = False

    table.configure(yscrollcommand=y_scrollbar.set)
    table.configure(xscrollcommand=update_x_scrollbar)
    table.bind("<Configure>", lambda _event: update_x_scrollbar(*table.xview()), add="+")
    root.after_idle(lambda: update_x_scrollbar(*table.xview()))

    status_frame = ttk.Frame(root, padding=(10, 4))
    status_frame.grid(row=3, column=0, sticky="ew")
    status_frame.columnconfigure(0, weight=1)

    status_bar = ttk.Label(status_frame, textvariable=status_var, anchor="w")
    status_bar.grid(row=0, column=0, sticky="w")

    footer_info = ttk.Label(status_frame, text=APP_FOOTER_TEXT, anchor="e")
    footer_info.grid(row=0, column=1, sticky="e")

    settings_button = ttk.Label(status_frame, text="设置", cursor="hand2")
    settings_button.grid(row=0, column=2, sticky="e", padx=(8, 0))
    settings_button.bind("<Button-1>", lambda _event: open_settings_window())

    context_menu = tk.Menu(root, tearoff=0)

    def get_sort_key(value: str):
        text = normalize_value(value)
        try:
            return (0, Decimal(text))
        except InvalidOperation:
            return (1, text.lower())

    def refresh_sort_headings() -> None:
        for col in current_headers:
            suffix = ""
            if current_sort_column == col:
                suffix = " ↓" if current_sort_is_desc else " ↑"
            table.heading(col, text=f"{col}{suffix}", command=lambda c=col: sort_table_by_column(c))

    def sort_table_by_column(column_name: str) -> None:
        nonlocal current_sort_column, current_sort_is_desc
        if column_name not in current_headers:
            return

        col_index = current_headers.index(column_name)
        descending = sort_descending.get(column_name, False)

        item_ids = list(table.get_children(""))
        sorted_ids = sorted(
            item_ids,
            key=lambda item_id: get_sort_key(str(table.item(item_id, "values")[col_index])),
            reverse=descending,
        )

        for row_idx, item_id in enumerate(sorted_ids):
            table.move(item_id, "", row_idx)

        current_sort_column = column_name
        current_sort_is_desc = descending
        sort_descending[column_name] = not descending
        refresh_sort_headings()

    def rebuild_query_inputs(headers: List[str]) -> None:
        # 每次加载新表后按动态表头重建查询输入区
        local_query_headers = get_query_headers(headers)

        for widget in query_input_frame.winfo_children():
            widget.destroy()

        query_vars.clear()
        for col_idx, header in enumerate(local_query_headers):
            value_var = tk.StringVar()
            query_vars[header] = value_var
            ttk.Label(query_input_frame, text=f"{header}:").grid(row=0, column=col_idx * 2, sticky="w")
            ttk.Entry(query_input_frame, textvariable=value_var, width=QUERY_INPUT_WIDTH).grid(
                row=0,
                column=col_idx * 2 + 1,
                padx=(6, 12),
                sticky="w",
            )

        if not local_query_headers:
            if not headers:
                return
            ttk.Label(query_input_frame, text=f"当前表格无可查询字段（{KEY_COLUMN_A} 与 {KEY_COLUMN_B} 之间无列）").grid(
                row=0,
                column=0,
                sticky="w",
            )

    def load_type_mapping_from_file() -> Tuple[Dict[str, Dict[str, str]], Optional[str]]:
        # 兼容旧配置格式：{"名称": "excel路径"}
        nonlocal config_path
        if config_path is None:
            config_path = get_preferred_config_path()
        return load_type_mapping_with_status(config_path)

    def get_ui_config_path() -> Path:
        nonlocal config_path
        if config_path is None:
            config_path = get_preferred_config_path()
        return config_path.with_name(UI_CONFIG_FILE)

    def load_app_title_from_file() -> str:
        ui_config_path = get_ui_config_path()
        if not ui_config_path.exists():
            return DEFAULT_MAIN_TITLE

        try:
            payload = json.loads(ui_config_path.read_text(encoding="utf-8"))
            if isinstance(payload, dict):
                title = str(payload.get("title", DEFAULT_MAIN_TITLE)).strip()
                return title or DEFAULT_MAIN_TITLE
        except (json.JSONDecodeError, OSError, TypeError, ValueError):
            pass
        return DEFAULT_MAIN_TITLE

    def save_app_title_to_file(title_text: str) -> bool:
        ui_config_path = get_ui_config_path()
        payload = {"title": title_text}
        try:
            ui_config_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            return True
        except OSError as exc:
            messagebox.showerror("保存失败", f"标题保存失败: {exc}")
            return False

    def save_type_mapping_to_file(mapping_to_save: Dict[str, Dict[str, str]]) -> bool:
        nonlocal config_path
        if config_path is None:
            config_path = get_preferred_config_path()
        payload = dump_signed_type_mapping(mapping_to_save)
        try:
            config_path.write_text(payload, encoding="utf-8")
            return True
        except OSError:
            # 程序目录不可写时，回退到用户目录保存。
            fallback_path = get_user_config_path()
            try:
                fallback_path.parent.mkdir(parents=True, exist_ok=True)
                fallback_path.write_text(payload, encoding="utf-8")
                config_path = fallback_path
                messagebox.showwarning(
                    "配置路径已切换",
                    f"当前目录不可写，已切换到用户目录保存配置:\n{fallback_path}",
                )
                return True
            except OSError as exc:
                messagebox.showerror("保存失败", f"配置保存失败: {exc}")
                return False

    def get_type_excel_path(name: str) -> str:
        return type_mapping.get(name, {}).get("excel", "").strip()

    def get_type_icon_path(name: str) -> str:
        return type_mapping.get(name, {}).get("icon", "").strip()

    def create_sized_icon(path_text: str, target_size: int = TYPE_ICON_TARGET_SIZE) -> Optional[Any]:
        path = Path(path_text)
        if not path.exists():
            return None

        try:
            cache_path = str(path.resolve())
        except (OSError, RuntimeError, ValueError):
            cache_path = str(path)

        cache_key = (cache_path, target_size)
        cached_image = icon_cache.get(cache_key)
        if cached_image is not None:
            return cached_image

        # 优先用 Pillow：缩放质量更好，尺寸控制更精确。
        try:
            from PIL import Image, ImageTk

            with Image.open(path) as pil_image:
                if pil_image.mode not in ("RGB", "RGBA"):
                    pil_image = pil_image.convert("RGBA")

                width, height = pil_image.size
                if width <= 0 or height <= 0:
                    return None

                # 固定显示框：无论原图大小，都按目标尺寸等比缩放（可放大/可缩小）。
                scale = min(target_size / width, target_size / height)
                resized_w = max(1, int(round(width * scale)))
                resized_h = max(1, int(round(height * scale)))
                resized = pil_image.resize((resized_w, resized_h), Image.Resampling.LANCZOS)
            image = ImageTk.PhotoImage(resized)
            icon_cache[cache_key] = image
            return image
        except (ImportError, OSError, ValueError):
            pass

        try:
            image = tk.PhotoImage(file=str(path))
        except tk.TclError:
            return None

        width = image.width()
        height = image.height()
        # 回退到 Tk：只能整数倍率缩放，但也能保证不超过目标尺寸。
        if max(width, height) > target_size:
            scale = get_icon_subsample_scale(width, height, target_size)
            if scale <= 1:
                icon_cache[cache_key] = image
                return image
            try:
                image = image.subsample(scale, scale)
            except tk.TclError:
                return None
        icon_cache[cache_key] = image
        return image

    def apply_settings_button_icon() -> None:
        nonlocal settings_button_icon_image
        if not options_icon_path.exists():
            return

        footer_height = max(footer_info.winfo_reqheight(), footer_info.winfo_height(), 1)
        target_height = max(int(round(footer_height * 0.6)), 1)

        try:
            from PIL import Image, ImageTk

            with Image.open(options_icon_path) as pil_image:
                if pil_image.mode not in ("RGB", "RGBA"):
                    pil_image = pil_image.convert("RGBA")

                width, height = pil_image.size
                if width <= 0 or height <= 0:
                    return

                resized_height = max(target_height, 1)
                resized_width = max(1, int(round(width * (resized_height / height))))
                resized = pil_image.resize((resized_width, resized_height), Image.Resampling.LANCZOS)

            settings_button_icon_image = ImageTk.PhotoImage(resized)
            settings_button.configure(text="", image=settings_button_icon_image, compound="center")
            return
        except (ImportError, OSError, ValueError, tk.TclError):
            pass

        try:
            image = tk.PhotoImage(file=str(options_icon_path))
            width = image.width()
            height = image.height()
            if width <= 0 or height <= 0:
                return

            if height > target_height:
                scale = get_icon_subsample_scale(width, height, target_height)
                if scale > 1:
                    image = image.subsample(scale, scale)

            settings_button_icon_image = image
            settings_button.configure(text="", image=settings_button_icon_image, compound="center")
        except tk.TclError:
            return

    def update_type_icon(name: Optional[str] = None) -> None:
        nonlocal type_icon_image
        # 默认按“已加载类型”显示图标，而不是当前下拉框选中值。
        # 这样用户必须点“加载”后，图标才代表真实生效类型。
        selected_name = name if name is not None else loaded_type_name
        icon_path_text = get_type_icon_path(selected_name) or default_type_icon_path
        image = create_sized_icon(icon_path_text)
        type_icon_canvas.delete("all")
        if image is None:
            type_icon_image = None
            type_icon_canvas.create_text(TYPE_ICON_TARGET_SIZE // 2, TYPE_ICON_TARGET_SIZE // 2, text="[icon]")
            return

        type_icon_image = image
        type_icon_canvas.create_image(TYPE_ICON_TARGET_SIZE // 2, TYPE_ICON_TARGET_SIZE // 2, image=type_icon_image)

    def schedule_type_icon_update(name: Optional[str] = None, delay_ms: int = 0) -> None:
        nonlocal pending_icon_update_id

        if pending_icon_update_id is not None:
            try:
                root.after_cancel(pending_icon_update_id)
            except tk.TclError:
                pass

        def run_icon_update() -> None:
            nonlocal pending_icon_update_id
            pending_icon_update_id = None
            update_type_icon(name)

        if delay_ms > 0:
            pending_icon_update_id = root.after(delay_ms, run_icon_update)
        else:
            pending_icon_update_id = root.after_idle(run_icon_update)

    def refresh_type_options(preferred_name: Optional[str] = None, delay_icon_ms: int = 0) -> None:
        nonlocal loaded_type_name
        names = list(type_mapping.keys())
        type_combo["values"] = names

        if preferred_name and preferred_name in type_mapping:
            type_var.set(preferred_name)
        else:
            current_name = type_var.get().strip()
            if current_name in type_mapping:
                pass
            elif names:
                type_var.set(names[0])
            else:
                type_var.set("")

        if loaded_type_name and loaded_type_name not in type_mapping:
            loaded_type_name = ""

        schedule_type_icon_update(loaded_type_name, delay_ms=delay_icon_ms)

    def update_table_headers(headers: List[str]) -> None:
        nonlocal current_headers, current_sort_column, current_sort_is_desc
        current_headers = get_display_headers(headers)
        sort_descending.clear()
        current_sort_column = None
        current_sort_is_desc = False
        table.configure(columns=tuple(current_headers))
        for col in current_headers:
            table.column(col, width=TABLE_COLUMN_WIDTH, minwidth=40, anchor="center")
        refresh_sort_headings()

    def auto_fit_table_columns(
        rows: List[Dict[str, str]],
        sample_limit: Optional[int] = None,
        only_expand: bool = False,
    ) -> None:
        if not current_headers:
            return

        rows_to_measure = rows
        if sample_limit is not None and len(rows) > sample_limit:
            rows_to_measure = rows[:sample_limit]

        horizontal_padding = 28
        for col in current_headers:
            max_pixel_width = table_measure_font.measure(str(col))
            for row in rows_to_measure:
                cell_text = str(row.get(col, ""))
                max_pixel_width = max(max_pixel_width, table_measure_font.measure(cell_text))

            target_width = max(max_pixel_width + horizontal_padding, 40)
            if only_expand:
                current_width = int(table.column(col, "width"))
                table.column(col, width=max(target_width, current_width))
            else:
                table.column(col, width=target_width)

    def clear_table() -> None:
        for item_id in table.get_children():
            table.delete(item_id)

    def render_rows(rows: List[Dict[str, str]]) -> None:
        nonlocal current_sort_column, current_sort_is_desc, pending_full_autofit_id, render_generation
        # New result sets should start unsorted to avoid stale sort indicators.
        current_sort_column = None
        current_sort_is_desc = False
        sort_descending.clear()
        refresh_sort_headings()
        render_generation += 1

        if pending_full_autofit_id is not None:
            try:
                root.after_cancel(pending_full_autofit_id)
            except tk.TclError:
                pass
            pending_full_autofit_id = None

        clear_table()
        for row in rows:
            table.insert("", "end", values=tuple(row.get(col, "") for col in current_headers))
        auto_fit_table_columns(rows, sample_limit=AUTO_FIT_SAMPLE_LIMIT)

        if len(rows) > AUTO_FIT_SAMPLE_LIMIT:
            generation_snapshot = render_generation

            def run_full_autofit() -> None:
                nonlocal pending_full_autofit_id
                pending_full_autofit_id = None
                if generation_snapshot != render_generation:
                    return
                auto_fit_table_columns(rows, only_expand=True)

            pending_full_autofit_id = root.after_idle(run_full_autofit)

    def get_selected_row_values() -> Optional[List[str]]:
        selected = table.selection()
        if not selected:
            messagebox.showwarning("未选择记录", "请先在结果表中选择一条记录。")
            return None

        values = table.item(selected[0], "values")
        return list(values)

    def copy_text(text: str, label: str) -> None:
        root.clipboard_clear()
        root.clipboard_append(text)
        root.update()
        status_var.set(f"已复制{label}: {text}")

    def copy_cell_by_column(column_index: int, header_name: str) -> None:
        row_values = get_selected_row_values()
        if row_values is None:
            return

        if column_index < 0 or column_index >= len(row_values):
            messagebox.showwarning("复制失败", f"无法定位列“{header_name}”。")
            return

        copy_text(row_values[column_index], header_name)

    def show_context_menu(event: tk.Event) -> None:
        # 根据鼠标所在列动态生成“复制<列名>”右键菜单
        item_id = table.identify_row(event.y)
        column_id = table.identify_column(event.x)
        if item_id:
            table.selection_set(item_id)
            table.focus(item_id)

        if not table.selection():
            return

        if not column_id or not column_id.startswith("#"):
            return

        try:
            column_index = int(column_id[1:]) - 1
        except ValueError:
            return

        if column_index < 0 or column_index >= len(current_headers):
            return

        header_name = current_headers[column_index]

        context_menu.delete(0, "end")
        context_menu.add_command(
            label=f"复制{header_name}",
            command=lambda idx=column_index, name=header_name: copy_cell_by_column(idx, name),
        )

        try:
            context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()

    def get_selected_excel_path() -> Optional[Path]:
        selected_type = type_var.get().strip()
        if not selected_type:
            messagebox.showwarning("未选择类型", "请先选择一个类型。")
            return None

        file_path = get_type_excel_path(selected_type)
        if not file_path:
            messagebox.showwarning("类型未配置", "当前类型未绑定 Excel 文件，请在设置中配置。")
            return None

        return Path(file_path)

    def load_data() -> bool:
        # 加载后同步刷新：表头、输入项、结果区
        nonlocal records, query_indexes, loaded_type_name, fuzzy_query_fields, normalized_query_rows
        path = get_selected_excel_path()
        if path is None:
            return False

        selected_type = type_var.get().strip()
        try:
            headers, records = load_records(path)
            update_table_headers(headers)
            rebuild_query_inputs(headers)
            fuzzy_query_fields = get_fuzzy_query_fields(headers)
            # 只为“可查询字段”建立索引，避免无用内存开销。
            query_headers = get_query_headers(headers)
            exact_query_headers = [header for header in query_headers if header not in fuzzy_query_fields]
            query_indexes = build_query_indexes(records, exact_query_headers)
            normalized_query_rows = build_normalized_query_rows(records, query_headers)
            render_rows(records)
            loaded_type_name = selected_type
            update_type_icon(loaded_type_name)
            status_var.set(f"已加载类型“{selected_type}”，共 {len(records)} 条记录")
            return True
        except (FileNotFoundError, PermissionError, ValueError, OSError) as exc:
            records = []
            fuzzy_query_fields = set()
            normalized_query_rows = []
            status_var.set("加载失败，请检查设置与文件。")
            messagebox.showerror("加载失败", str(exc))
            return False
        except (ModuleNotFoundError, ImportError):
            records = []
            fuzzy_query_fields = set()
            normalized_query_rows = []
            status_var.set("加载失败：缺少 openpyxl 依赖。")
            messagebox.showerror("缺少依赖", "未安装 openpyxl，请先安装后再加载。")
            return False

    def load_data_from_path(path: Path, source_label: str) -> bool:
        # 允许通过命令行参数直接指定 Excel，并在 GUI 启动后自动加载。
        nonlocal records, query_indexes, loaded_type_name, fuzzy_query_fields, normalized_query_rows
        try:
            headers, records = load_records(path)
            update_table_headers(headers)
            rebuild_query_inputs(headers)
            fuzzy_query_fields = get_fuzzy_query_fields(headers)
            query_headers = get_query_headers(headers)
            exact_query_headers = [header for header in query_headers if header not in fuzzy_query_fields]
            query_indexes = build_query_indexes(records, exact_query_headers)
            normalized_query_rows = build_normalized_query_rows(records, query_headers)
            render_rows(records)
            loaded_type_name = ""
            update_type_icon(loaded_type_name)
            status_var.set(f"已通过{source_label}加载数据，共 {len(records)} 条记录")
            return True
        except (FileNotFoundError, PermissionError, ValueError, OSError) as exc:
            records = []
            fuzzy_query_fields = set()
            normalized_query_rows = []
            status_var.set("加载失败，请检查设置与文件。")
            messagebox.showerror("加载失败", str(exc))
            return False
        except (ModuleNotFoundError, ImportError):
            records = []
            fuzzy_query_fields = set()
            normalized_query_rows = []
            status_var.set("加载失败：缺少 openpyxl 依赖。")
            messagebox.showerror("缺少依赖", "未安装 openpyxl，请先安装后再加载。")
            return False

    def on_type_selected(_event: tk.Event) -> None:
        nonlocal records, query_indexes, fuzzy_query_fields, normalized_query_rows
        # 切换类型时清空旧结果，防止用户误把旧数据当成新类型数据。
        records = []
        query_indexes = {}
        fuzzy_query_fields = set()
        normalized_query_rows = []
        clear_table()
        for value_var in query_vars.values():
            value_var.set("")
        selected_type = type_var.get().strip()
        if selected_type:
            status_var.set(f"当前类型: {selected_type}，点击“加载”读取数据。")

    def open_settings_window() -> None:
        password = simpledialog.askstring(
            "口令验证",
            "请输入设置口令：",
            show="*",
            parent=root,
        )
        if password is None:
            status_var.set("已取消打开设置。")
            return
        if password != SETTINGS_PASSCODE:
            status_var.set("口令错误，无法打开设置。")
            messagebox.showerror("口令错误", "口令不正确，无法打开设置。")
            return

        settings_window = tk.Toplevel(root)
        settings_window.title("设置")
        settings_width = SETTINGS_WINDOW_WIDTH
        settings_height = SETTINGS_WINDOW_HEIGHT
        settings_window.geometry(f"{settings_width}x{settings_height}")
        settings_window.minsize(720, 420)
        settings_window.transient(root)
        settings_window.grab_set()
        settings_window.update_idletasks()

        root_x = root.winfo_rootx()
        root_y = root.winfo_rooty()
        root_w = root.winfo_width()
        root_h = root.winfo_height()

        pos_x = root_x + max((root_w - settings_width) // 2, 0)
        pos_y = root_y + max((root_h - settings_height) // 2, 0)
        settings_window.geometry(f"{settings_width}x{settings_height}+{pos_x}+{pos_y}")
        settings_window.lift()
        settings_window.focus_force()

        settings_window.columnconfigure(0, weight=1)
        settings_window.rowconfigure(0, weight=1)

        container = ttk.Frame(settings_window, padding=10)
        container.grid(row=0, column=0, sticky="nsew")
        container.columnconfigure(0, weight=1)
        container.columnconfigure(1, weight=0)
        container.rowconfigure(1, weight=1)

        form_frame = ttk.LabelFrame(container, text="编辑区", padding=(10, 8))
        form_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        form_frame.columnconfigure(1, weight=0)
        form_frame.columnconfigure(3, weight=2)

        setting_name_var = tk.StringVar()
        setting_path_var = tk.StringVar()
        setting_icon_var = tk.StringVar()
        setting_title_var = tk.StringVar(value=app_title_var.get())
        draft_type_mapping = {
            key: {"excel": value.get("excel", ""), "icon": value.get("icon", "")}
            for key, value in type_mapping.items()
        }
        pending_loaded_type_name = loaded_type_name
        has_unsaved_changes = False

        def mark_settings_dirty(changed: bool = True) -> None:
            nonlocal has_unsaved_changes
            has_unsaved_changes = changed

        original_title_text = app_title_var.get()

        def on_title_changed(*_args) -> None:
            title_now = setting_title_var.get().strip() or DEFAULT_MAIN_TITLE
            mapping_changed = draft_type_mapping != type_mapping
            mark_settings_dirty(mapping_changed or title_now != original_title_text)

        setting_title_var.trace_add("write", on_title_changed)

        ttk.Label(form_frame, text="标题:").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=(8, 0))
        ttk.Entry(form_frame, textvariable=setting_title_var).grid(row=0, column=1, columnspan=3, sticky="ew", pady=(8, 0))

        ttk.Label(form_frame, text="保存后生效").grid(row=0, column=4, padx=(8, 0), sticky="w", pady=(8, 0))

        ttk.Label(form_frame, text="名称:").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=(8, 0))
        ttk.Entry(form_frame, textvariable=setting_name_var, width=18).grid(row=1, column=1, sticky="w", pady=(8, 0))
        ttk.Label(form_frame, text="Excel 文件:").grid(row=1, column=2, sticky="w", padx=(4, 8), pady=(8, 0))
        ttk.Entry(form_frame, textvariable=setting_path_var).grid(row=1, column=3, sticky="ew", pady=(8, 0))

        def browse_setting_excel() -> None:
            filename = filedialog.askopenfilename(
                title="选择 Excel 文件",
                filetypes=[("Excel 文件", "*.xlsx *.xlsm *.xltx *.xltm"), ("所有文件", "*.*")],
            )
            if filename:
                setting_path_var.set(filename)

        ttk.Button(form_frame, text="浏览...", command=browse_setting_excel).grid(row=1, column=4, padx=(8, 0), pady=(8, 0))

        ttk.Label(form_frame, text="图标路径:").grid(row=2, column=0, sticky="w", padx=(0, 8), pady=(8, 0))
        ttk.Entry(form_frame, textvariable=setting_icon_var).grid(row=2, column=1, columnspan=3, sticky="ew", pady=(8, 0))

        def browse_setting_icon() -> None:
            filename = filedialog.askopenfilename(
                title="选择图标文件",
                filetypes=[("图像文件", "*.png *.gif *.ppm *.pgm"), ("所有文件", "*.*")],
            )
            if filename:
                setting_icon_var.set(filename)

        ttk.Button(form_frame, text="选择图标", command=browse_setting_icon).grid(row=2, column=4, padx=(8, 0), pady=(8, 0))

        setting_table_frame = ttk.LabelFrame(container, text="类型列表", padding=(0, 6, 0, 0))
        setting_table_frame.grid(row=1, column=0, columnspan=2, sticky="nsew")
        setting_table_frame.rowconfigure(0, weight=1)
        setting_table_frame.columnconfigure(0, weight=1)

        setting_columns = ("name", "path", "icon")
        setting_table = ttk.Treeview(setting_table_frame, columns=setting_columns, show="headings", height=10)
        setting_table.heading("name", text="名称")
        setting_table.heading("path", text="Excel 文件路径")
        setting_table.heading("icon", text="图标路径")
        setting_table.column("name", width=170, anchor="center")
        setting_table.column("path", width=320, anchor="w")
        setting_table.column("icon", width=220, anchor="w")
        setting_table.grid(row=0, column=0, sticky="nsew")

        setting_scroll = ttk.Scrollbar(setting_table_frame, orient="vertical", command=setting_table.yview)
        setting_scroll.grid(row=0, column=1, sticky="ns")
        setting_x_scroll = ttk.Scrollbar(setting_table_frame, orient="horizontal", command=setting_table.xview)
        setting_x_scroll.grid(row=1, column=0, sticky="ew")
        setting_table.configure(yscrollcommand=setting_scroll.set)
        setting_table.configure(xscrollcommand=setting_x_scroll.set)

        action_frame = ttk.Frame(container)
        action_frame.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(10, 0))
        action_frame.columnconfigure(6, weight=1)

        def refresh_setting_table(select_name: Optional[str] = None) -> None:
            for item_id in setting_table.get_children():
                setting_table.delete(item_id)

            for name, config in draft_type_mapping.items():
                row_id = setting_table.insert(
                    "",
                    "end",
                    values=(name, config.get("excel", ""), config.get("icon", "")),
                )
                if select_name and name == select_name:
                    setting_table.selection_set(row_id)
                    setting_table.focus(row_id)

        def on_setting_row_select(_event: tk.Event) -> None:
            selected = setting_table.selection()
            if not selected:
                return
            values = setting_table.item(selected[0], "values")
            setting_name_var.set(str(values[0]))
            setting_path_var.set(str(values[1]))
            setting_icon_var.set(str(values[2]))

        def save_setting() -> None:
            nonlocal pending_loaded_type_name
            name = setting_name_var.get().strip()
            excel = setting_path_var.get().strip()
            icon = setting_icon_var.get().strip()
            save_action = "新增"

            if not name:
                messagebox.showwarning("输入不完整", "请填写名称。", parent=settings_window)
                return
            if not excel:
                messagebox.showwarning("输入不完整", "请填写或选择 Excel 文件路径。", parent=settings_window)
                return

            selected = setting_table.selection()
            new_mapping = {
                key: {"excel": value.get("excel", ""), "icon": value.get("icon", "")}
                for key, value in draft_type_mapping.items()
            }
            if selected:
                old_name = str(setting_table.item(selected[0], "values")[0]).strip()
                # 仅当“选中名称”和“输入名称”一致时，按更新处理。
                # 否则按新增处理，避免连续新增时误覆盖刚添加的项。
                if old_name == name:
                    save_action = "更新"
                    if old_name in new_mapping:
                        del new_mapping[old_name]
                    if pending_loaded_type_name == old_name:
                        pending_loaded_type_name = name
                elif name in new_mapping:
                    messagebox.showwarning("名称重复", "该名称已存在，请使用其他名称。", parent=settings_window)
                    return
            else:
                if name in new_mapping:
                    messagebox.showwarning("名称重复", "该名称已存在，请使用其他名称。", parent=settings_window)
                    return

            new_mapping[name] = {"excel": excel, "icon": icon}
            draft_type_mapping.clear()
            draft_type_mapping.update(new_mapping)
            mark_settings_dirty()
            refresh_setting_table(name)
            status_var.set(f"已{save_action}类型“{name}”（待保存）")

        def delete_setting() -> None:
            nonlocal pending_loaded_type_name
            selected = setting_table.selection()
            if not selected:
                messagebox.showwarning("未选择记录", "请先选择要删除的类型。", parent=settings_window)
                return

            name = str(setting_table.item(selected[0], "values")[0]).strip()
            if not messagebox.askyesno("确认删除", f"确定删除类型“{name}”吗？", parent=settings_window):
                return

            new_mapping = {
                key: {"excel": value.get("excel", ""), "icon": value.get("icon", "")}
                for key, value in draft_type_mapping.items()
            }
            if name in new_mapping:
                del new_mapping[name]
            if pending_loaded_type_name == name:
                pending_loaded_type_name = ""
            draft_type_mapping.clear()
            draft_type_mapping.update(new_mapping)
            mark_settings_dirty()
            refresh_setting_table()
            status_var.set(f"已删除类型“{name}”（待保存）")

        def move_setting(direction: int) -> None:
            selected = setting_table.selection()
            if not selected:
                messagebox.showwarning("未选择记录", "请先选择要移动的类型。", parent=settings_window)
                return

            selected_name = str(setting_table.item(selected[0], "values")[0]).strip()
            ordered_names = list(draft_type_mapping.keys())
            try:
                current_index = ordered_names.index(selected_name)
            except ValueError:
                return

            target_index = current_index + direction
            if target_index < 0 or target_index >= len(ordered_names):
                return

            ordered_names[current_index], ordered_names[target_index] = (
                ordered_names[target_index],
                ordered_names[current_index],
            )

            reordered_mapping = {
                name: {
                    "excel": draft_type_mapping.get(name, {}).get("excel", ""),
                    "icon": draft_type_mapping.get(name, {}).get("icon", ""),
                }
                for name in ordered_names
            }

            draft_type_mapping.clear()
            draft_type_mapping.update(reordered_mapping)
            mark_settings_dirty()
            refresh_setting_table(select_name=selected_name)
            move_action = "上移" if direction < 0 else "下移"
            status_var.set(f"已{move_action}类型“{selected_name}”（待保存）")

        def save_all_settings() -> None:
            nonlocal loaded_type_name, original_title_text

            title_text = setting_title_var.get().strip() or DEFAULT_MAIN_TITLE
            previous_title_text = load_app_title_from_file()

            def save_title() -> bool:
                return save_app_title_to_file(title_text)

            def save_mapping() -> bool:
                return save_type_mapping_to_file(draft_type_mapping)

            def rollback_title() -> bool:
                rolled_back = save_app_title_to_file(previous_title_text)
                if not rolled_back:
                    messagebox.showwarning(
                        "回滚失败",
                        "类型配置保存失败，且标题回滚失败，请手动检查标题配置。",
                        parent=settings_window,
                    )
                return rolled_back

            if not save_settings_atomically(save_title, save_mapping, rollback_title):
                return

            type_mapping.clear()
            type_mapping.update(draft_type_mapping)
            loaded_type_name = pending_loaded_type_name
            app_title_var.set(title_text)
            original_title_text = title_text
            refresh_type_options(type_var.get().strip() if type_var.get().strip() in type_mapping else None)
            mark_settings_dirty(False)
            status_var.set("设置已保存")

        def close_settings_window() -> None:
            if has_unsaved_changes:
                should_close = messagebox.askyesno(
                    "未保存更改",
                    "检测到未保存的配置更改，确定不保存并关闭吗？",
                    parent=settings_window,
                )
                if not should_close:
                    return
            settings_window.destroy()

        ttk.Button(action_frame, text="新增/更新", command=save_setting).grid(row=0, column=0)
        ttk.Button(action_frame, text="删除", command=delete_setting).grid(row=0, column=1, padx=(8, 0))
        ttk.Button(action_frame, text="上移", command=lambda: move_setting(-1)).grid(row=0, column=2, padx=(8, 0))
        ttk.Button(action_frame, text="下移", command=lambda: move_setting(1)).grid(row=0, column=3, padx=(8, 0))
        ttk.Button(action_frame, text="保存", command=save_all_settings).grid(row=0, column=7, padx=(8, 0), sticky="e")
        ttk.Button(action_frame, text="关闭", command=close_settings_window).grid(row=0, column=8, padx=(8, 0), sticky="e")

        setting_table.bind("<<TreeviewSelect>>", on_setting_row_select)
        settings_window.protocol("WM_DELETE_WINDOW", close_settings_window)
        refresh_setting_table()

    def run_query() -> None:
        # 如果还没加载数据，按当前类型先自动加载一次。
        if not records and not load_data():
            return

        conditions = {
            header: value_var.get().strip()
            for header, value_var in query_vars.items()
            if value_var.get().strip()
        }

        if not conditions:
            messagebox.showwarning("输入不完整", "请至少输入一个查询条件。")
            return

        candidate_indexes = get_candidate_indexes(query_indexes, conditions)
        results = query_records(
            records,
            conditions,
            candidate_indexes=candidate_indexes,
            fuzzy_fields=fuzzy_query_fields,
            normalized_rows=normalized_query_rows,
        )
        render_rows(results)

        if results:
            status_var.set(f"查询完成: 共 {len(results)} 条匹配结果")
        else:
            status_var.set("查询完成: 未找到匹配结果")

    def clear_query_inputs() -> None:
        # 仅清空查询输入，不影响已加载的数据。
        for value_var in query_vars.values():
            value_var.set("")

    ttk.Button(type_control_frame, text="加载", command=load_data).grid(row=1, column=2, sticky="w", padx=(8, 0))
    ttk.Button(query_frame, text="清除", command=clear_query_inputs).grid(row=0, column=1, padx=(8, 0), sticky="w")
    ttk.Button(query_frame, text="查询", command=run_query, style="Query.TButton").grid(row=0, column=2, padx=(8, 0), sticky="w")

    table.bind("<Button-3>", show_context_menu)
    type_combo.bind("<<ComboboxSelected>>", on_type_selected)

    rebuild_query_inputs(current_headers)

    root.bind("<Return>", lambda _event: run_query())

    def initialize_type_mapping() -> None:
        loaded_mapping, config_warning = load_type_mapping_from_file()
        type_mapping.update(loaded_mapping)
        if config_warning:
            messagebox.showwarning("配置加载提醒", config_warning)
        app_title_var.set(load_app_title_from_file())
        refresh_type_options(delay_icon_ms=DEFERRED_STARTUP_ICON_DELAY_MS)

        # 若启动时尝试过命令行参数加载（无论成功或失败），
        # 保留 load_data_from_path 给出的状态提示，不被异步初始化覆盖。
        if startup_load_attempted or records:
            return

        if type_var.get().strip():
            status_var.set(f"当前类型: {type_var.get().strip()}，点击“加载”读取数据。")
        else:
            status_var.set("暂无类型，请点击“设置”新增。")

    if prefer_default_excel:
        startup_load_attempted = True
        load_data_from_path(default_excel, "命令行参数")

    root.after_idle(apply_window_icon)
    root.after_idle(apply_settings_button_icon)
    root.after_idle(initialize_type_mapping)

    root.mainloop()



def main() -> None:
    # 命令行入口：兼容旧参数并支持动态字段查询
    parser = argparse.ArgumentParser(description="根据直径和长度查询杠铃汇总数据")
    parser.add_argument(
        "--excel",
        default="杠铃汇总.xlsx",
        help="Excel 文件路径，默认: 杠铃汇总.xlsx",
    )
    parser.add_argument("--diameter", help="直径（可单独查询）")
    parser.add_argument("--length", help="长度（可单独查询）")
    parser.add_argument("--field", action="append", help="查询字段名（可重复）")
    parser.add_argument("--value", action="append", help="查询字段值（可重复）")
    parser.add_argument(
        "--cli",
        action="store_true",
        help="使用命令行交互模式（默认启动可视化窗口）",
    )
    parser.add_argument(
        "--self-check",
        action="store_true",
        help="运行开发自检（前导0、小数查询、损坏配置回退）",
    )
    args = parser.parse_args()
    excel_argument_provided = args.excel != parser.get_default("excel")

    if args.self_check:
        if not run_self_check():
            raise RuntimeError("自检未通过")
        return

    excel_path = Path(args.excel)

    dynamic_fields = args.field or []
    dynamic_values = args.value or []
    query_conditions: Optional[Dict[str, str]] = None

    if dynamic_fields or dynamic_values:
        if len(dynamic_fields) != len(dynamic_values):
            raise ValueError("--field 与 --value 的数量必须一致。")
        query_conditions = {key: val for key, val in zip(dynamic_fields, dynamic_values)}
    elif args.diameter is not None or args.length is not None:
        query_conditions = {"直径": args.diameter or "", "长度": args.length or ""}

    # 优先级：
    # 1) 传了明确查询条件 -> 直接命令行查询
    # 2) 指定 --cli -> 进入命令行交互
    # 3) 默认 -> 打开 GUI
    if query_conditions is not None or args.cli:
        try:
            get_openpyxl_symbols()
        except (ModuleNotFoundError, ImportError):
            raise RuntimeError("缺少依赖 openpyxl，请先执行: pip install -r requirements.txt")

    if query_conditions is not None:
        headers, records = load_records(excel_path)
        fuzzy_fields = get_fuzzy_query_fields(headers)
        results = query_records(records, query_conditions, fuzzy_fields=fuzzy_fields)
        print_results(headers, results)
    elif args.cli:
        headers, records = load_records(excel_path)
        run_interactive(headers, records)
    else:
        launch_gui(excel_path, prefer_default_excel=excel_argument_provided)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"程序执行失败: {exc}")
        raise SystemExit(1)
