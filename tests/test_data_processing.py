import json
import tempfile
import unittest
from pathlib import Path
from unittest import mock
import sys

try:
    from openpyxl import Workbook
except (ModuleNotFoundError, ImportError):
    Workbook = None

import retrieval_tool

from retrieval_tool import (
    build_query_indexes,
    dump_signed_type_mapping,
    get_fuzzy_query_fields,
    get_preferred_config_path,
    get_icon_subsample_scale,
    get_candidate_indexes,
    get_query_headers,
    load_records,
    load_type_mapping_from_file_path,
    load_type_mapping_with_status,
    normalize_value,
    query_records,
    save_settings_atomically,
)


class TestDataProcessing(unittest.TestCase):
    def test_save_settings_atomically_success(self):
        calls = []

        def save_title() -> bool:
            calls.append("title")
            return True

        def save_mapping() -> bool:
            calls.append("mapping")
            return True

        def rollback_title() -> bool:
            calls.append("rollback")
            return True

        self.assertTrue(save_settings_atomically(save_title, save_mapping, rollback_title))
        self.assertEqual(calls, ["title", "mapping"])

    def test_save_settings_atomically_fail_on_title(self):
        calls = []

        def save_title() -> bool:
            calls.append("title")
            return False

        def save_mapping() -> bool:
            calls.append("mapping")
            return True

        def rollback_title() -> bool:
            calls.append("rollback")
            return True

        self.assertFalse(save_settings_atomically(save_title, save_mapping, rollback_title))
        self.assertEqual(calls, ["title"])

    def test_save_settings_atomically_rolls_back_title(self):
        calls = []

        def save_title() -> bool:
            calls.append("title")
            return True

        def save_mapping() -> bool:
            calls.append("mapping")
            return False

        def rollback_title() -> bool:
            calls.append("rollback")
            return True

        self.assertFalse(save_settings_atomically(save_title, save_mapping, rollback_title))
        self.assertEqual(calls, ["title", "mapping", "rollback"])

    def test_get_query_headers(self):
        headers = ["序号", "/A", "直径", "长度", "/B"]
        self.assertEqual(get_query_headers(headers), ["直径", "长度"])

    def test_normalize_value_decimal_and_leading_zero(self):
        self.assertEqual(normalize_value("1200.50"), "1200.5")
        self.assertEqual(normalize_value("037332314020"), "037332314020")
        self.assertEqual(normalize_value("-0.0"), "0")

    def test_query_with_index_candidates(self):
        records = [
            {"直径": "20", "长度": "1200", "/A": "A"},
            {"直径": "20", "长度": "1500", "/A": "B"},
            {"直径": "22", "长度": "1200", "/A": "C"},
        ]
        indexes = build_query_indexes(records, ["直径", "长度"])
        candidates = get_candidate_indexes(indexes, {"直径": "20", "长度": "1200"})
        results = query_records(records, {"直径": "20", "长度": "1200"}, candidate_indexes=candidates)
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]["/A"], "A")

    def test_get_fuzzy_query_fields_from_key_ranges(self):
        headers = ["序号", "/A", "/C", "材质", "规格", "/D", "直径", "长度", "/B"]
        self.assertEqual(get_fuzzy_query_fields(headers), {"材质", "规格"})

    def test_get_fuzzy_query_fields_rejects_cross_ranges(self):
        headers = ["序号", "/A", "/C", "材质", "/E", "等级", "/D", "直径", "/F", "/B"]
        with self.assertRaisesRegex(ValueError, "不能交叉"):
            get_fuzzy_query_fields(headers)

    def test_query_records_supports_fuzzy_contains(self):
        records = [
            {"材质": "304不锈钢", "规格": "M10", "直径": "20"},
            {"材质": "碳钢", "规格": "M12", "直径": "20"},
        ]
        results = query_records(records, {"材质": "不锈"}, fuzzy_fields={"材质"})
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]["规格"], "M10")

    def test_get_icon_subsample_scale_boundary(self):
        self.assertEqual(get_icon_subsample_scale(120, 120, 120), 1)
        self.assertEqual(get_icon_subsample_scale(121, 121, 120), 2)
        self.assertEqual(get_icon_subsample_scale(239, 239, 120), 2)
        self.assertEqual(get_icon_subsample_scale(240, 240, 120), 2)

    @unittest.skipIf(Workbook is None, "openpyxl is not installed")
    def test_load_records_preserves_leading_zero_by_cell_format(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / "sample.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["序号", "/A", "直径", "长度", "/B"])
            ws.cell(row=2, column=1, value=1)
            code_cell = ws.cell(row=2, column=2, value=37332314020)
            code_cell.number_format = "000000000000"
            ws.cell(row=2, column=3, value=20)
            ws.cell(row=2, column=4, value=1200)
            ws.cell(row=2, column=5, value="项目A")
            wb.save(excel_path)
            wb.close()

            headers, records = load_records(excel_path)
            self.assertEqual(headers, ["序号", "/A", "直径", "长度", "/B"])
            self.assertEqual(records[0]["/A"], "037332314020")

    @unittest.skipIf(Workbook is None, "openpyxl is not installed")
    def test_load_records_rejects_fuzzy_keys_outside_main_range(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            excel_path = Path(tmp_dir) / "invalid_fuzzy_range.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["序号", "/C", "材质", "/D", "/A", "直径", "/B"])
            ws.append([1, "", "304", "", "A001", "20", "项目A"])
            wb.save(excel_path)
            wb.close()

            with self.assertRaisesRegex(ValueError, "标题位置错误"):
                load_records(excel_path)

    def test_config_loader_compat_and_fallback(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            config_path = Path(tmp_dir) / "types_config.json"

            # Old format compatibility
            config_path.write_text('{"默认": "C:/data/a.xlsx"}', encoding="utf-8")
            config = load_type_mapping_from_file_path(config_path)
            self.assertEqual(config["默认"]["excel"], "C:/data/a.xlsx")
            self.assertEqual(config["默认"]["icon"], "")

            # Corrupted config fallback
            config_path.write_text("{invalid_json", encoding="utf-8")
            config = load_type_mapping_from_file_path(config_path)
            self.assertEqual(config, {})

    def test_get_preferred_config_path_falls_back_to_user_dir(self):
        with tempfile.TemporaryDirectory() as app_dir, tempfile.TemporaryDirectory() as user_dir:
            app_base = Path(app_dir)
            user_base = Path(user_dir)

            with mock.patch.object(retrieval_tool, "get_app_base_dir", return_value=app_base), mock.patch.object(
                retrieval_tool,
                "get_user_config_dir",
                return_value=user_base,
            ):
                config_path = get_preferred_config_path()

            expected_path = user_base / "types_config.json"
            self.assertEqual(config_path, expected_path)
            self.assertTrue(config_path.exists())

            payload = json.loads(config_path.read_text(encoding="utf-8"))
            self.assertEqual(payload.get("_data"), {})
            self.assertTrue(payload.get("_signature"))

    def test_config_loader_rejects_partial_signed_fields(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            config_path = Path(tmp_dir) / "types_config.json"
            config_path.write_text('{"_data": {"默认": {"excel": "C:/a.xlsx", "icon": ""}}}', encoding="utf-8")
            config = load_type_mapping_from_file_path(config_path)
            self.assertEqual(config, {})

    def test_config_loader_accepts_signed_payload(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            config_path = Path(tmp_dir) / "types_config.json"
            mapping = {"默认": {"excel": "C:/a.xlsx", "icon": ""}}
            config_path.write_text(dump_signed_type_mapping(mapping), encoding="utf-8")
            config = load_type_mapping_from_file_path(config_path)
            self.assertEqual(config, mapping)

    def test_config_loader_rejects_unsigned_new_dict_format(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            config_path = Path(tmp_dir) / "types_config.json"
            config_path.write_text('{"默认": {"excel": "C:/a.xlsx", "icon": ""}}', encoding="utf-8")
            config = load_type_mapping_from_file_path(config_path)
            self.assertEqual(config, {})

    def test_config_loader_status_contains_warning_for_tampered_config(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            config_path = Path(tmp_dir) / "types_config.json"
            config_path.write_text('{"_data": {"默认": {"excel": "C:/a.xlsx", "icon": ""}}}', encoding="utf-8")
            mapping, warning = load_type_mapping_with_status(config_path)
            self.assertEqual(mapping, {})
            self.assertTrue(warning)
            self.assertIn("手动修改", warning)

    def test_main_uses_excel_arg_for_gui_preload(self):
        with mock.patch.object(retrieval_tool, "launch_gui") as launch_gui_mock:
            with mock.patch.object(sys, "argv", ["retrieval_tool.py", "--excel", "custom.xlsx"]):
                retrieval_tool.main()

        launch_gui_mock.assert_called_once_with(Path("custom.xlsx"), prefer_default_excel=True)

    def test_main_default_gui_does_not_force_preload(self):
        with mock.patch.object(retrieval_tool, "launch_gui") as launch_gui_mock:
            with mock.patch.object(sys, "argv", ["retrieval_tool.py"]):
                retrieval_tool.main()

        launch_gui_mock.assert_called_once_with(Path("杠铃汇总.xlsx"), prefer_default_excel=False)


if __name__ == "__main__":
    unittest.main()
